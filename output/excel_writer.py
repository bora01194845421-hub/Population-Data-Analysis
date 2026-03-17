"""
Excel 출력 모듈
- 시트 0: 요약 (KPI 카드)
- 시트 1: 지역별 인구 (Agent A) + 막대차트
- 시트 2: 행정동 인구 (Agent B) + 데이터바 조건부 서식
- 시트 3: 생애주기 분석 (Agent C) + 파이차트
- 시트 4: 행정동×생애주기 히트맵
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ───── 공통 색상 ─────
COLOR_HEADER_BG = "2E6DB4"  # 파란 계열 헤더
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEADER  = "D6E4F7"
COLOR_TOTAL_BG   = "FFF2CC"
COLOR_KPI_BG     = "EBF3FB"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_font(bold=True, color=COLOR_HEADER_FG, size=11):
    return Font(bold=bold, color=color, size=size)


def _header_fill(color=COLOR_HEADER_BG):
    return PatternFill("solid", fgColor=color)


def _apply_table_style(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    """헤더 행 서식 + 전체 테두리"""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=min_row, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in range(min_row + 1, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col > min_col else "left", vertical="center")


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    """DataFrame을 워크시트에 쓰고 (max_row, max_col) 반환"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
    return start_row + len(df), start_col + len(df.columns) - 1


def _autofit_columns(ws, min_col: int = 1, max_col: int | None = None):
    max_col = max_col or ws.max_column
    for col in range(min_col, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


# ════════════════════════════════════════════════════════════
# Sheet 0 — 요약 (KPI 카드)
# ════════════════════════════════════════════════════════════
def _sheet_summary(ws, region_df: pd.DataFrame, lifecycle_df: pd.DataFrame | None):
    ws.title = "0_요약"
    ws.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells("B2:H2")
    ws["B2"] = "경기도 인구자료 정리 서비스 — 요약"
    ws["B2"].font = title_font
    ws["B2"].alignment = Alignment(horizontal="center")

    kpi_fill = PatternFill("solid", fgColor=COLOR_KPI_BG)

    # ── 지역 KPI ──
    ws["B4"] = "▶ 지역별 인구 현황"
    ws["B4"].font = Font(bold=True, size=12)

    headers = ["지역", "총인구", "남자", "여자", "성비", "전국비율(%)"]
    for c_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=c_idx, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for r_idx, (_, row) in enumerate(region_df.iterrows(), start=6):
        for c_idx, col in enumerate(headers, start=2):
            val = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.fill = kpi_fill
            cell.alignment = Alignment(horizontal="right" if c_idx > 2 else "left")

    # ── 생애주기 KPI ──
    if lifecycle_df is not None and not lifecycle_df.empty:
        start_r = 6 + len(region_df) + 2
        ws.cell(row=start_r, column=2, value="▶ 수원시 생애주기 요약").font = Font(bold=True, size=12)
        lc_display = lifecycle_df[lifecycle_df["생애주기"] != "합계"]
        for c_idx, h in enumerate(["생애주기", "연령구간", "인구", "비율(%)"], start=2):
            cell = ws.cell(row=start_r + 1, column=c_idx, value=h)
            cell.font = _header_font()
            cell.fill = _header_fill()
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        for r_idx, (_, row) in enumerate(lc_display.iterrows(), start=start_r + 2):
            for c_idx, col in enumerate(["생애주기", "연령구간", "인구", "비율(%)"], start=2):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(col, ""))
                cell.border = THIN_BORDER
                cell.fill = kpi_fill
                cell.alignment = Alignment(horizontal="right" if c_idx > 3 else "left")

    _autofit_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet 1 — 지역별 인구 (Agent A) + 막대차트
# ════════════════════════════════════════════════════════════
def _sheet_region(ws, df: pd.DataFrame):
    ws.title = "1_지역별인구"

    max_row, max_col = _write_df(ws, df)
    _apply_table_style(ws, 1, max_row, 1, max_col)

    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 합계 행 강조
    total_fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
    for col in range(1, max_col + 1):
        ws.cell(row=max_row, column=col).fill = total_fill

    # ── 막대차트 (총인구) ──
    chart = BarChart()
    chart.type = "col"
    chart.title = "지역별 총인구"
    chart.y_axis.title = "인구 (명)"
    chart.x_axis.title = "지역"
    chart.style = 10
    chart.width = 20
    chart.height = 14

    # 총인구 컬럼 인덱스 찾기
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    pop_col_idx = next((i + 1 for i, h in enumerate(headers) if h == "총인구"), None)
    region_col_idx = 1

    if pop_col_idx:
        data_ref = Reference(ws, min_col=pop_col_idx, min_row=1, max_row=max_row - 1)
        cats_ref = Reference(ws, min_col=region_col_idx, min_row=2, max_row=max_row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{max_row + 3}")

    _autofit_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet 2 — 행정동 인구 (Agent B) + 데이터바
# ════════════════════════════════════════════════════════════
def _sheet_dong(ws, detail: pd.DataFrame, gu_summary: pd.DataFrame):
    ws.title = "2_행정동인구"

    # 구별 소계 먼저
    ws["A1"] = "【구별 소계】"
    ws["A1"].font = Font(bold=True, size=12)

    max_row_gu, max_col_gu = _write_df(ws, gu_summary, start_row=2)
    _apply_table_style(ws, 2, max_row_gu, 1, max_col_gu)
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col_gu)}{max_row_gu}"

    gap = 3
    detail_start = max_row_gu + gap

    ws.cell(row=detail_start, column=1, value="【행정동 전체 목록】").font = Font(bold=True, size=12)
    max_row_d, max_col_d = _write_df(ws, detail, start_row=detail_start + 1)
    _apply_table_style(ws, detail_start + 1, max_row_d, 1, max_col_d)
    ws.auto_filter.ref = f"A{detail_start + 1}:{get_column_letter(max_col_d)}{max_row_d}"

    # 총인구 컬럼에 DataBar 조건부 서식
    headers = [ws.cell(row=detail_start + 1, column=c).value for c in range(1, max_col_d + 1)]
    pop_col = next((i + 1 for i, h in enumerate(headers) if h == "총인구"), None)
    if pop_col:
        col_letter = get_column_letter(pop_col)
        ws.conditional_formatting.add(
            f"{col_letter}{detail_start + 2}:{col_letter}{max_row_d}",
            DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color="4472C4"),
        )

    _autofit_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet 3 — 생애주기 분석 (Agent C) + 파이차트
# ════════════════════════════════════════════════════════════
def _sheet_lifecycle(ws, lifecycle_df: pd.DataFrame, dependency_df: pd.DataFrame):
    ws.title = "3_생애주기분석"

    ws["A1"] = "【생애주기별 인구】"
    ws["A1"].font = Font(bold=True, size=12)

    max_row_lc, max_col_lc = _write_df(ws, lifecycle_df, start_row=2)
    _apply_table_style(ws, 2, max_row_lc, 1, max_col_lc)

    # 합계 행 강조
    total_fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
    for col in range(1, max_col_lc + 1):
        ws.cell(row=max_row_lc, column=col).fill = total_fill
        ws.cell(row=max_row_lc, column=col).font = Font(bold=True)

    # ── 파이차트 ──
    chart = PieChart()
    chart.title = "수원시 생애주기 인구 비율"
    chart.style = 10
    chart.width = 18
    chart.height = 14

    lc_no_total = lifecycle_df[lifecycle_df["생애주기"] != "합계"]
    stage_col = 1
    pop_col = next((i + 1 for i, h in enumerate(lifecycle_df.columns) if h == "인구"), None)

    if pop_col:
        data_ref = Reference(ws, min_col=pop_col, min_row=3, max_row=max_row_lc - 1)
        cats_ref = Reference(ws, min_col=stage_col, min_row=3, max_row=max_row_lc - 1)
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{max_row_lc + 3}")

    # ── 부양비 지표 ──
    dep_start = max_row_lc + 3 + 22
    ws.cell(row=dep_start, column=1, value="【부양비 지표】").font = Font(bold=True, size=12)
    max_row_dep, max_col_dep = _write_df(ws, dependency_df, start_row=dep_start + 1)
    _apply_table_style(ws, dep_start + 1, max_row_dep, 1, max_col_dep)

    # 비율 컬럼에 ColorScale
    for col in range(2, max_col_dep + 1):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{dep_start + 2}:{col_letter}{max_row_dep}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    _autofit_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet 4 — 행정동×생애주기 히트맵
# ════════════════════════════════════════════════════════════
def _sheet_heatmap(ws, dong_pivot: pd.DataFrame):
    ws.title = "4_행정동생애주기"

    if dong_pivot.empty:
        ws["A1"] = "행정동별 연령 데이터가 없습니다."
        return

    ws["A1"] = "【행정동별 생애주기 비율(%) 히트맵】"
    ws["A1"].font = Font(bold=True, size=12)

    max_row, max_col = _write_df(ws, dong_pivot, start_row=2)
    _apply_table_style(ws, 2, max_row, 1, max_col)
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"

    # 생애주기 컬럼에 각각 ColorScale
    stage_names = ["영유아", "아동", "청소년", "청년", "중장년", "장년", "노년기"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]
    for stage in stage_names:
        if stage in headers:
            col_idx = headers.index(stage) + 1
            col_letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}3:{col_letter}{max_row}",
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="2E6DB4",
                ),
            )

    _autofit_columns(ws)


# ════════════════════════════════════════════════════════════
# 메인 진입점
# ════════════════════════════════════════════════════════════
def save(
    output_path: Union[str, Path],
    region_df: pd.DataFrame,
    agent_b_results: dict,
    agent_c_results: dict,
) -> Path:
    """
    Parameters
    ----------
    output_path     : 저장할 Excel 파일 경로
    region_df       : Agent A 결과
    agent_b_results : Agent B 결과 dict
    agent_c_results : Agent C 결과 dict

    Returns
    -------
    저장된 파일 경로
    """
    wb = Workbook()
    wb.remove(wb.active)

    lifecycle_df = agent_c_results.get("lifecycle", pd.DataFrame())
    dependency_df = agent_c_results.get("dependency", pd.DataFrame())
    dong_pivot = agent_c_results.get("dong_pivot", pd.DataFrame())
    detail = agent_b_results.get("detail", pd.DataFrame())
    gu_summary = agent_b_results.get("gu_summary", pd.DataFrame())

    ws0 = wb.create_sheet("0_요약")
    _sheet_summary(ws0, region_df, lifecycle_df if not lifecycle_df.empty else None)

    ws1 = wb.create_sheet("1_지역별인구")
    _sheet_region(ws1, region_df)

    ws2 = wb.create_sheet("2_행정동인구")
    _sheet_dong(ws2, detail, gu_summary)

    ws3 = wb.create_sheet("3_생애주기분석")
    _sheet_lifecycle(ws3, lifecycle_df, dependency_df)

    ws4 = wb.create_sheet("4_행정동생애주기")
    _sheet_heatmap(ws4, dong_pivot)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def to_bytes(
    region_df: pd.DataFrame,
    agent_b_results: dict,
    agent_c_results: dict,
) -> bytes:
    """Streamlit download_button용 인메모리 바이트 반환"""
    buf = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    lifecycle_df = agent_c_results.get("lifecycle", pd.DataFrame())
    dependency_df = agent_c_results.get("dependency", pd.DataFrame())
    dong_pivot = agent_c_results.get("dong_pivot", pd.DataFrame())
    detail = agent_b_results.get("detail", pd.DataFrame())
    gu_summary = agent_b_results.get("gu_summary", pd.DataFrame())

    ws0 = wb.create_sheet("0_요약")
    _sheet_summary(ws0, region_df, lifecycle_df if not lifecycle_df.empty else None)
    ws1 = wb.create_sheet("1_지역별인구")
    _sheet_region(ws1, region_df)
    ws2 = wb.create_sheet("2_행정동인구")
    _sheet_dong(ws2, detail, gu_summary)
    ws3 = wb.create_sheet("3_생애주기분석")
    _sheet_lifecycle(ws3, lifecycle_df, dependency_df)
    ws4 = wb.create_sheet("4_행정동생애주기")
    _sheet_heatmap(ws4, dong_pivot)

    wb.save(buf)
    return buf.getvalue()
